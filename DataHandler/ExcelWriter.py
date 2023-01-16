import openpyxl
import openpyxl.styles
import openpyxl.utils
from openpyxl.styles import Border


EXCEL_REGISTER_COLORS = {
    "DATA": "92D050",
    "COVER": "1F497D"
}

QUESTION_TYPES_COLORS = {
    1: {"color": "FFFFFF",
        "description": "Single Choice"},
    2: {"color": "92D050",
        "description": "Multiple Choice"
        },
    3: {"color": "1F497D",
        "description": "Custom User Choice"},
    4: {"color": "FFFFFF",
        "description": "Single Choice"},
}


class ExcelWriter:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

        self.worksheet.title = "Cover"
        self.customize_sheet(self.worksheet, EXCEL_REGISTER_COLORS["COVER"])

        self.create_sheet("Data")
        self.customize_sheet(self.workbook["Data"], EXCEL_REGISTER_COLORS["DATA"])

    def create_sheet(self, name):
        return self.workbook.create_sheet(name)

    def write(self, data):

        for row in data:
            self.worksheet.append(row)

    def save(self):
        self.workbook.save(self.filename)

    def create_cover(self, data, color_translation: bool = True):
        self.worksheet = self.workbook["Cover"]
        offset_top = 3

        question_merge_cells_range = 3
        question_cell = 1
        question_merge_end_column = question_cell + question_merge_cells_range - 1

        answer_merge_cells_range = 4
        answer_cell = question_merge_end_column + 1
        answer_merge_end_column = answer_cell + answer_merge_cells_range - 1

        outer_side_style = openpyxl.styles.Side(border_style="medium", color="000000")
        inner_side_style = openpyxl.styles.Side(border_style="thin", color="000000")

        column = 1
        row = offset_top + 1

        def set_border(upper_left_cell, lower_right_cell, side_style=outer_side_style):
            if lower_right_cell==upper_left_cell:
                self.worksheet[upper_left_cell].border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)

            upper_left_cell_border = Border(left=side_style, top=side_style)
            upper_right_cell_border = Border(right=side_style, top=side_style)
            lower_left_cell_border = Border(left=side_style, bottom=side_style)
            lower_right_cell_border = Border(right=side_style, bottom=side_style)
            upper_cell_border = Border(top=side_style)
            lower_cell_border = Border(bottom=side_style)
            left_cell_border = Border(left=side_style)
            right_cell_border = Border(right=side_style)

            # set upper border
            for column in range(upper_left_cell.column, lower_right_cell.column + 1):
                self.worksheet.cell(row=upper_left_cell.row, column=column).border = upper_cell_border

            # set lower border
            for column in range(upper_left_cell.column, lower_right_cell.column + 1):
                self.worksheet.cell(row=lower_right_cell.row, column=column).border = lower_cell_border

            # set left border
            for row in range(upper_left_cell.row, lower_right_cell.row + 1):
                self.worksheet.cell(row=row, column=upper_left_cell.column).border = left_cell_border

            # set right border
            for row in range(upper_left_cell.row, lower_right_cell.row + 1):
                self.worksheet.cell(row=row, column=lower_right_cell.column).border = right_cell_border

            self.worksheet[upper_left_cell.coordinate].border = upper_left_cell_border
            self.worksheet[lower_right_cell.coordinate].border = lower_right_cell_border
            self.worksheet.cell(row=upper_left_cell.row, column=lower_right_cell.column).border = upper_right_cell_border
            self.worksheet.cell(row=lower_right_cell.row, column=upper_left_cell.column).border = lower_left_cell_border

        def inner_border(upper_left_cell, lower_right_cell, side_style=inner_side_style):

            for row in range(upper_left_cell.row, lower_right_cell.row + 1):
                for column in range(upper_left_cell.column, lower_right_cell.column + 1):
                    self.worksheet.cell(row=row, column=column).border = Border(left=side_style, right=side_style, top=side_style, bottom=side_style)


        def color_explanation(color_code_explanation_column=15,start_row=row,width=2,height_description=2,height_color=2):


            if isinstance(color_code_explanation_column, str):
                color_code_explanation_column = openpyxl.utils.column_index_from_string(color_code_explanation_column)


            width = width-1
            height_description = height_description-1
            height_color = height_color-1

            self.worksheet.merge_cells(start_row=start_row, start_column=color_code_explanation_column, end_row=start_row+height_description, end_column=color_code_explanation_column+width)
            self.worksheet.cell(row=start_row, column=color_code_explanation_column).value = "Question Type"
            self.worksheet.cell(row=start_row, column=color_code_explanation_column).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            set_border(self.worksheet.cell(row=start_row, column=color_code_explanation_column), self.worksheet.cell(row=start_row+height_description, column=color_code_explanation_column+width))

            self.worksheet.merge_cells(start_row=start_row+height_description+1, start_column=color_code_explanation_column, end_row=start_row+height_description+1+height_color, end_column=color_code_explanation_column+width)
            self.worksheet.cell(row=start_row+height_description+1, column=color_code_explanation_column).value = "Color Code"
            self.worksheet.cell(row=start_row+height_description+1, column=color_code_explanation_column).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            set_border(self.worksheet.cell(row=start_row+height_description+1, column=color_code_explanation_column), self.worksheet.cell(row=start_row+height_description+1+height_color, column=color_code_explanation_column+width))

            self.worksheet.cell(row=start_row, column=color_code_explanation_column).fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            self.worksheet.cell(row=start_row+height_description+1, column=color_code_explanation_column).fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")




            color_code_explanation_column+=width+1

            for color_code in QUESTION_TYPES_COLORS.keys():
                if color_code==4:
                    continue
                description_row_start = start_row
                description_row_end = description_row_start + height_description
                color_row_start = description_row_end + 1
                color_row_end = color_row_start + height_color

                self.worksheet.cell(row=description_row_start, column=color_code_explanation_column).value = QUESTION_TYPES_COLORS[color_code]["description"]
                self.worksheet.merge_cells(start_row=description_row_start, start_column=color_code_explanation_column, end_row=description_row_end, end_column=color_code_explanation_column + width)
                set_border(self.worksheet.cell(row=description_row_start, column=color_code_explanation_column), self.worksheet.cell(row=description_row_end, column=color_code_explanation_column + width))

                self.worksheet.cell(row=color_row_start, column=color_code_explanation_column).fill = openpyxl.styles.PatternFill("solid", fgColor=QUESTION_TYPES_COLORS[color_code]["color"])
                self.worksheet.cell(row=color_row_start, column=color_code_explanation_column).value = f'#{QUESTION_TYPES_COLORS[color_code]["color"]}'
                self.worksheet.merge_cells(start_row=color_row_start, start_column=color_code_explanation_column, end_row=color_row_end, end_column=color_code_explanation_column + width)
                set_border(self.worksheet.cell(row=color_row_start, column=color_code_explanation_column), self.worksheet.cell(row=color_row_end, column=color_code_explanation_column + width))


                self.worksheet.cell(row=row, column=color_code_explanation_column).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
                self.worksheet.cell(row=color_row_start, column=color_code_explanation_column).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

                color_code_explanation_column += width+1





        color_explanation(color_code_explanation_column="K", start_row=4)
        for x in data.keys():
            top_row = row

            answer_options_count = len(data[x]["answer_options"])

            cell = self.worksheet.cell(row, question_cell)
            cell.value = data[x]["question_text"]#x
            cell.font = openpyxl.styles.Font(bold=True, size=12)
            self.worksheet.merge_cells(start_row=row, start_column=question_cell, end_row=row + answer_options_count, end_column=question_merge_end_column)

            self.worksheet.cell(row, column).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")
            set_border(self.worksheet.cell(row, column), self.worksheet.cell(row + answer_options_count, question_merge_end_column))
            if color_translation:
                if data[x]["question_type"] in QUESTION_TYPES_COLORS.keys():
                    self.worksheet.cell(row, column).fill = openpyxl.styles.PatternFill("solid", fgColor=QUESTION_TYPES_COLORS[data[x]["question_type"]]["color"])

            for y in range(answer_options_count):
                self.worksheet.cell(row=row, column=answer_cell).value = data[x]["answer_options"][y]

                self.worksheet.cell(row=row, column=answer_cell).font = openpyxl.styles.Font(size=12)
                self.worksheet.merge_cells(start_row=row, start_column=question_cell, end_row=row, end_column=question_merge_end_column)
                self.worksheet.merge_cells(start_row=row, start_column=answer_cell, end_row=row, end_column=answer_merge_end_column)
                row += 1

            self.worksheet.merge_cells(start_row=row, start_column=answer_cell, end_row=row, end_column=answer_merge_end_column)
            row += 1

            set_border(self.worksheet.cell(top_row, answer_cell), self.worksheet.cell(row - 1, answer_merge_end_column))

    # fancy features-----------------------------------------------------------------------------------------------------

    def customize_sheet(self, sheet, color):
        sheet.sheet_properties.tabColor = color

    def rotation(self, rotation, cell):
        self.worksheet[cell].alignment = openpyxl.styles.Alignment(textRotation=rotation)

    def column_width(self, width, column):
        self.worksheet.column_dimensions[column].width = width

    def row_height(self, height, row):
        self.worksheet.row_dimensions[row].height = height

    def cell_font(self, font, cell):
        self.worksheet[cell].font = font

    def cell_left_align(self, cell):
        self.worksheet[cell].alignment = openpyxl.styles.Alignment(horizontal="left")

    def cell_right_align(self, cell):
        self.worksheet[cell].alignment = openpyxl.styles.Alignment(horizontal="right")

    def cell_center_align(self, cell):
        self.worksheet[cell].alignment = openpyxl.styles.Alignment(horizontal="center")

    def create_data(self, data, header=None):
        self.worksheet = self.workbook["Data"]
        row = 1

        if header:
            for k,v in header.items():
                self.worksheet.cell(row=row, column=int(k)).value = v["question_text"]
                self.worksheet.cell(row=row, column=int(k)).font = openpyxl.styles.Font(bold=True)
                # fill alignment
                self.worksheet.cell(row=row, column=int(k)).alignment = openpyxl.styles.Alignment(horizontal="fill")
                # freeze header
            self.worksheet.freeze_panes = self.worksheet.cell(row=2, column=1)


        for user in data.keys():
            column = 0
            for question in data[user].keys():
                string = f"{data[user][question]}"
                string = string.replace(",", ";")
                string = string.replace("[", "")
                string = string.replace("]", "")
                self.worksheet.cell(row=row + 1, column=int(question)).value = f"[{string}]"
                column += 1
            row += 1
